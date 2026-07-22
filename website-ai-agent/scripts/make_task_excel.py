"""Generate the dogfood_tester task-tracking Excel workbook on the Desktop.

Two sheets: remaining engineering work, and actions required from the project owner. One-off
utility, not part of the shipped package.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path.home() / "Desktop" / "dogfood_tester(excel-task).xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (No., Category, Task, Detail, Priority, Effort, Blocked by, Status)
REMAINING = [
    (
        "Eval harness",
        "Add load / latency benchmark scenario",
        "Larger dynamic fixture run to measure steps-per-page, latency, tokens under size.",
        "Low",
        "M",
        "-",
        "Not started",
    ),
    (
        "Eval fixtures",
        "Add authenticated-flow fixture",
        "Exercise storage-state recipes and scoped login flows without real credentials.",
        "Medium",
        "M",
        "-",
        "Not started",
    ),
    (
        "LLM testing",
        "Record and commit role LLM replay cassettes",
        "Cassette infra exists; record real-model responses so prompt regressions gate in CI.",
        "Medium",
        "M",
        "Owner: API key",
        "Blocked",
    ),
    (
        "Scaling",
        "Add Postgres checkpoint backend",
        "Alternative to SQLite for multi-worker deployments; state interface already allows it.",
        "Low",
        "L",
        "Owner: Postgres instance",
        "Blocked",
    ),
    (
        "Accessibility",
        "Full ARIA accessible-name computation",
        "Current accessible-name is a pragmatic ARIA subset; make it spec-complete.",
        "Low",
        "L",
        "-",
        "Not started",
    ),
    (
        "Vision",
        "Multimodal visual QA checks",
        "vision/ module is heuristic and off by default; add real multimodal checks.",
        "Low",
        "L",
        "Owner: multimodal key",
        "Blocked",
    ),
]

# (No., Action, Why it is needed, When needed, Priority)
REQUIRED = [
    (
        "Provide an LLM endpoint or API key",
        "To run the agent against real models and to record replay cassettes. Free option: run "
        "Ollama locally and set WA_LLM__BASE_URL=http://localhost:11434/v1 (zero cost).",
        "Before any real run",
        "High",
    ),
    (
        "Authorize the target site(s)",
        "Confirm which sites the agent may explore. Point it only at sites you own or staging; "
        "the agent runs a real browser and must be authorized.",
        "Before running on a real site",
        "High",
    ),
    (
        "Configure PyPI Trusted Publishing",
        "release.yml publishes on v* tags via OIDC. Create the PyPI project and a 'pypi' GitHub "
        "environment for the publish job to activate; otherwise it stays a no-op.",
        "Before first PyPI release",
        "Medium",
    ),
    (
        "Decide build priorities",
        "Choose which Remaining Work items to build next (eval fixtures and the 3 detectors are "
        "the highest-value). Reply with the numbers.",
        "To continue development",
        "High",
    ),
    (
        "Review the destructive-action policy",
        "Default is safe-explore (no destructive actions). If you want the agent to submit forms "
        "or exercise mutating actions on staging, enable it explicitly and confirm scope.",
        "Before test-mode runs that mutate",
        "Medium",
    ),
    (
        "Provide a Postgres instance (optional)",
        "Only if multi-worker horizontal scaling is wanted; single-worker SQLite works today.",
        "Only for scaling",
        "Low",
    ),
    (
        "Set CI secrets for hosted-model eval (optional)",
        "If you want nightly eval against a hosted model in CI, add the provider key as a repo "
        "secret; default CI runs keyless with scripted models.",
        "Only for live-model CI",
        "Low",
    ),
]


def _style_header(ws, headers: list[str], row: int) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER


def _finalize(ws, widths: list[int], header_row: int, last_row: int, ncols: int) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(header_row + 1, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = CENTER if c == 1 else WRAP
            cell.border = BORDER
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.sheet_view.showGridLines = False


def build() -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Remaining Work"
    ws1["A1"] = "dogfood_tester - Remaining Engineering Work"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:H1")
    headers1 = ["No.", "Category", "Task", "Detail", "Priority", "Effort", "Blocked by", "Status"]
    _style_header(ws1, headers1, 3)
    for i, row in enumerate(REMAINING, start=1):
        ws1.append([]) if False else None
        ws1.cell(row=3 + i, column=1, value=i)
        for c, val in enumerate(row, start=2):
            ws1.cell(row=3 + i, column=c, value=val)
    _finalize(ws1, [6, 16, 34, 52, 10, 8, 18, 13], 3, 3 + len(REMAINING), 8)

    ws2 = wb.create_sheet("Required From You")
    ws2["A1"] = "dogfood_tester - Actions Required From You"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:E1")
    headers2 = ["No.", "Action required", "Why it is needed", "When needed", "Priority"]
    _style_header(ws2, headers2, 3)
    for i, row in enumerate(REQUIRED, start=1):
        ws2.cell(row=3 + i, column=1, value=i)
        for c, val in enumerate(row, start=2):
            ws2.cell(row=3 + i, column=c, value=val)
    _finalize(ws2, [6, 34, 60, 26, 10], 3, 3 + len(REQUIRED), 5)

    wb.save(OUT)
    print(f"saved: {OUT}")
    print(f"Remaining Work rows: {len(REMAINING)} | Required From You rows: {len(REQUIRED)}")


if __name__ == "__main__":
    build()
